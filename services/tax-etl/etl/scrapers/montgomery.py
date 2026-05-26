"""
Montgomery County — North Houston growth corridor.
Sources:
  - Montgomery County delinquent tax roll XLSX (direct download, 75MB, 524K rows)
"""
import logging
import httpx
import io
import re
from typing import List, Dict
from openpyxl import load_workbook
from bs4 import BeautifulSoup

HEADERS = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}
TAX_FORMS_URL = "https://www.mctotx.org/property/property_tax_forms.php"


def get_latest_xlsx_url() -> str:
    """Get the latest delinquent tax roll XLSX URL from the forms page."""
    resp = httpx.get(TAX_FORMS_URL, headers=HEADERS, follow_redirects=True, timeout=30)
    soup = BeautifulSoup(resp.text, "lxml")
    for a in soup.find_all("a", href=True):
        href = a["href"]
        text = a.get_text(strip=True)
        if "delinquent" in text.lower() and "detail" in text.lower() and ".xlsx" in href.lower():
            if href.startswith("http"):
                return href
            prefix = "/" if not href.startswith("/") else ""
            return f"https://www.mctotx.org{prefix}{href}"
    return ""


def download_and_parse_xlsx(url: str) -> List[Dict]:
    """Download and parse the delinquent tax roll XLSX."""
    try:
        with httpx.stream("GET", url, headers=HEADERS, follow_redirects=True, timeout=600) as resp:
            resp.raise_for_status()
            content = b""
            for chunk in resp.iter_bytes(chunk_size=65536):
                content += chunk

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        records = {}
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or i == 1:
                continue
            if not headers:
                headers = [str(c or f"col_{j}").lower().strip() for j, c in enumerate(row)]
                continue
            row_dict = dict(zip(headers, row))

            can = str(row_dict.get("can", "") or "").strip().lstrip("0")
            aprdistacc = str(row_dict.get("aprdistacc", "") or "").strip()
            pid = aprdistacc if aprdistacc else can
            if not pid or pid == "None":
                continue

            addr = str(row_dict.get("addrstring", "") or "").strip()
            pnumber = str(row_dict.get("pnumber", "") or "").strip()
            pstname = str(row_dict.get("pstname", "") or "").strip()
            property_addr = f"{pnumber} {pstname}".strip()

            try:
                levy = float(row_dict.get("levy_balance", 0) or 0)
                tot_percan = float(row_dict.get("tot_percan", 0) or 0)
            except (ValueError, TypeError):
                levy = 0
                tot_percan = 0

            year = str(row_dict.get("year", "") or "")

            if pid not in records:
                records[pid] = {
                    "parcel_id": pid,
                    "owner_name": addr.split(",")[0].strip() if addr else "",
                    "address": property_addr or addr,
                    "appraised_value": 0,
                    "est_tax_due": 0,
                    "source": "montgomery_xlsx",
                    "county": "montgomery",
                    "is_delinquent": True,
                    "on_auction_list": False,
                    "_years": set(),
                    "_total_due": 0,
                }
            records[pid]["_years"].add(year)
            records[pid]["_total_due"] += levy + tot_percan

        # Finalize totals
        result = []
        for pid, r in records.items():
            r["est_tax_due"] = round(r["_total_due"], 2)
            r["tax_years_due"] = ", ".join(sorted(r["_years"]))
            del r["_years"], r["_total_due"]
            result.append(r)

        logging.info(f"[montgomery/xlsx] {len(result):,} delinquent parcels ({len(records):,} unique)")
        return result
    except Exception as e:
        logging.error(f"[montgomery/xlsx] failed: {e}")
        return []


def ingest() -> List[Dict]:
    xlsx_url = get_latest_xlsx_url()
    if xlsx_url:
        logging.info(f"[montgomery] downloading XLSX from {xlsx_url}")
        records = download_and_parse_xlsx(xlsx_url)
    else:
        logging.warning("[montgomery] no XLSX URL found")
        records = []
    seen = {}
    for r in records:
        pid = r.get("parcel_id") or r.get("address", "")
        if pid not in seen:
            seen[pid] = r
    return list(seen.values())
