"""
Brazoria County — SE Houston / Gulf Coast.
Sources:
  - PBFCM tax sale PDF (primary, confirmed working)
  - Brazoria County delinquent tax roll (OneDrive Excel)
  - LGBS taxsales.lgbs.com (backup)
"""
import logging
import httpx
import io
import re
from typing import List, Dict
from openpyxl import load_workbook
from etl.scrapers.pbfcm import ingest_county as pbfcm_ingest

HEADERS = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}


def scrape_delinquent_roll() -> List[Dict]:
    """Download Brazoria County delinquent tax roll from OneDrive."""
    url = "https://1drv.ms/x/c/c5b1985dca72f232/IQALb5drpwCETI730d2prW5DAfgL-YxAvVPQbGfc5c1Mjc4?e=WmY3TE"
    try:
        resp = httpx.get(url, headers=HEADERS, follow_redirects=True, timeout=120)
        resp.raise_for_status()
        wb = load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
        ws = wb.active
        records = []
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c).lower().strip() if c else f"col_{j}" for j, c in enumerate(row)]
                continue
            row_dict = dict(zip(headers, [str(c or "") for c in row]))
            pid = row_dict.get("account", row_dict.get("parcel", row_dict.get("col_0", "")))
            if not pid or pid == "None":
                continue
            records.append({
                "parcel_id": pid.strip(),
                "owner_name": row_dict.get("owner", row_dict.get("owner_name", "")).strip(),
                "address": row_dict.get("property_address", row_dict.get("situs", "")).strip(),
                "est_tax_due": float(re.sub(r"[^\d.]", "", row_dict.get("amount_due", row_dict.get("total_due", "0"))) or 0),
                "source": "brazoria_delinquent",
                "county": "brazoria",
                "is_delinquent": True,
                "on_auction_list": False,
            })
        logging.info(f"[brazoria/delinquent] {len(records)} records from OneDrive")
        return records
    except Exception as e:
        logging.warning(f"[brazoria/delinquent] failed: {e}")
        return []


def ingest() -> List[Dict]:
    records = []
    records.extend(pbfcm_ingest("brazoria"))
    records.extend(scrape_delinquent_roll())
    seen = {}
    for r in records:
        pid = r.get("parcel_id") or r.get("address", "")
        if pid not in seen or r.get("on_auction_list"):
            seen[pid] = r
    return list(seen.values())
